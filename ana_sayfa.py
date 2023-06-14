import cv2
import face_recognition
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.uic import loadUi
from PyQt5.QtCore import pyqtSlot, QTimer, QDate
from PyQt5.QtWidgets import QApplication,QDialog,QMessageBox
import numpy as np
import datetime
import os
from openpyxl import load_workbook

def sayi_to_harf(sayi):
    harf = chr(sayi + 96)
    return (harf)

class Ui_OutputDialog(QDialog):
    def __init__(self):
        super(Ui_OutputDialog, self).__init__()
        loadUi("./ana_sayfa.ui", self)

        now = QDate.currentDate()
        current_date = now.toString('dd MMMM yyyy dddd')
        current_time = datetime.datetime.now().strftime("%H:%M:%S")
        self.Date_Label.setText(current_date)
        self.Time_Label.setText(current_time)
        self.image = None

        self.CikisButon.clicked.connect(self.cikis_buton_clicked)
        self.YenileButon.clicked.connect(self.yenile_buton_clicked)

    @pyqtSlot()
    def yenile_buton_clicked(self):
        self.OdaLabel.clear()
        self.NameLabel.clear()
        self.StatusLabel.clear()
        self.HoursLabel.clear()
        self.MinLabel.clear()

    def cikis_buton_clicked(self):
        QApplication.quit()

    def startVideo(self, camera_name):  # threading
        self.timer = QTimer(self)
        if len(camera_name) == 1:
            self.capture = cv2.VideoCapture(int(camera_name))
        else:
            self.capture = cv2.VideoCapture(camera_name)
        for img in images:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            boxes = face_recognition.face_locations(img)
            encodes_cur_frame = face_recognition.face_encodings(img, boxes)[0]

            self.encode_list.append(encodes_cur_frame)
        self.timer.timeout.connect(self.guncelle)
        self.timer.start(10)

    def egit(self):
        path = 'Ogrenciler'
        if not os.path.exists(path):
            os.mkdir(path)
        global images
        images = []  # Ogrencilerin isim listesi
        self.class_names = []
        self.encode_list = []
        self.TimeList1 = []
        self.TimeList2 = []
        attendance_list = os.listdir(path)
        for cl in attendance_list:
            cur_img = cv2.imread(f'{path}/{cl}')
            images.append(cur_img)
            self.class_names.append(os.path.splitext(cl)[0])

    def face_rec_(self, frame, encode_list_known, class_names):

        def yuzEslestir(tcKimlik):
            wb = load_workbook("Pansiyon_Ogrenci_Listesi.xlsx")
            sayfa = wb.worksheets[0]
            ws = wb.active

            def search(aranan):
                for satir in range(1, sayfa.max_row):
                    for sutun in range(1, sayfa.max_column):
                        if sayfa.cell(row=satir, column=sutun).value == aranan:
                            return (satir, sutun)

            if self.OnaylaButon.isChecked():

                self.OnaylaButon.setEnabled(False)

                if (tcKimlik != 'Yüz Eşleştirilemedi'):
                            satir,sutun = search(int(tcKimlik))
                            sutunad = sayi_to_harf(sutun+2)
                            oda = sayi_to_harf(sutun+1)
                            adSoyad = ws[sutunad.upper() + str(satir)].value
                            odaNo = ws[oda.upper() + str(satir)].value
                            cevap = QMessageBox.question(self, 'Hoşgeldiniz ' + adSoyad, ' Yoklamanızı Onaylıyor musunuz?' ,
                                                               QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                            if cevap == QMessageBox.Yes:

                                satir,sutun = search(int(tcKimlik))
                                self.OnaylaButon.setChecked(False)
                                self.NameLabel.setText(adSoyad)
                                self.OdaLabel.setText(odaNo)
                                sutunsaat = sayi_to_harf(sutun+3)
                                sutuntarih = sayi_to_harf(sutun+4)
                                print(sutunsaat,sutuntarih)
                                print(type(sutunsaat), type(sutuntarih))
                                ws[sutunsaat.upper()+str(satir)]=datetime.datetime.now().strftime("%H:%M:%S")
                                ws[sutuntarih.upper() + str(satir)] = datetime.datetime.now().strftime("%d/%m/%y")
                                wb.save("Pansiyon_Ogrenci_Listesi.xlsx")
                                self.StatusLabel.setText('Yoklama Alındı')
                                self.HoursLabel.setText(datetime.datetime.now().strftime("%H:%M:%S"))
                                self.MinLabel.setText('')
                                self.Time1 = datetime.datetime.now()
                                self.OnaylaButon.setEnabled(True)
                            elif cevap == QMessageBox.No:
                                self.StatusLabel.setText('Yoklama Onaylanmadı')
                                self.OnaylaButon.setEnabled(False)
                            else:
                                print('Onaylanmadı')


        faces_cur_frame = face_recognition.face_locations(frame)
        encodes_cur_frame = face_recognition.face_encodings(frame, faces_cur_frame)

        for encodeFace, faceLoc in zip(encodes_cur_frame, faces_cur_frame):
            match = face_recognition.compare_faces(encode_list_known, encodeFace, tolerance=0.50)
            face_dis = face_recognition.face_distance(encode_list_known, encodeFace)
            tcKimlik = "Yüz Eşleştirilemedi"
            best_match_index = np.argmin(face_dis)

            if match[best_match_index]:
                tcKimlik = class_names[best_match_index].upper()
                y1, x2, y2, x1 = faceLoc
                cv2.rectangle(frame, (x1-10, y1-10), (x2+10, y2+10), (0, 0, 255), 2)
                cv2.rectangle(frame, (x1, y2 - 20), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(frame, tcKimlik, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 1)
            yuzEslestir(tcKimlik)

        return frame


    def guncelle(self):
        ret, self.image = self.capture.read()
        self.displayImage(self.image, self.encode_list, self.class_names, 1)

    def displayImage(self, image, encode_list, class_names, window=1):

        image = cv2.resize(image, (640, 480))
        try:
            image = self.face_rec_(image, encode_list, class_names)
        except Exception as e:
            print(e)
        qformat = QImage.Format_Indexed8
        if len(image.shape) == 3:
            if image.shape[2] == 4:
                qformat = QImage.Format_RGBA8888
            else:
                qformat = QImage.Format_RGB888
        outImage = QImage(image, image.shape[1], image.shape[0], image.strides[0], qformat)
        outImage = outImage.rgbSwapped()

        if window == 1:
            self.imgLabel.setPixmap(QPixmap.fromImage(outImage))
            self.imgLabel.setScaledContents(True)

